import sys
import os
sys.path.append(os.getcwd())

from openpyxl import load_workbook
import json
import platform
os_name = platform.system()

import pandas as pd


ConfigDataPath = os.path.join(os.getcwd(),'utils','InfluxDB','ConfigData')
Opts = { # In the xlsx file there are the list of all the sensors/quantities of interest linked to the NodeID name in OPCUA 
        'xlsx_InputFileName_Readings'   : 'Variabili_Database_Influx.xlsx', 
        'json_OutputFileName_Readings'  : 'DBdataConfig.json',
        'ConfigDataPath'                : ConfigDataPath, }

def ReadDBVarsFromExcel(Opts):
    if os_name != 'Windows':
        raise Exception('This function is only available for Windows OS')
    

    xlsxFileName = os.path.join(Opts['ConfigDataPath'],Opts['xlsx_InputFileName_Readings'])
    excel_file = load_workbook(xlsxFileName)

    
    sheet_names = excel_file.sheetnames

    sheet_names = [sheet for sheet in sheet_names if sheet != 'Data']
    DB_data = {}
    DB_NodeIDtoTag = {}
    DB_TagtoNodeID = {}
    for sheet in sheet_names: 
        DB_data[str(sheet)] = {}
        
        excel_table = pd.read_excel(
            xlsxFileName,
            sheet_name=sheet,
            index_col=False,
            engine='openpyxl')
        
        excel_table.fillna('-', inplace = True)

        rows = excel_table.index

        for row in rows:
            if sheet == 'Meter':
                # In the excel file, the FRER name is not correct and must be replaced
                NodeID      = 'ns=4;s=' + excel_table['NodeID'][row]
                FRERName    = str(excel_table['nomeFRER'][row])
                NodeID      = NodeID.replace('nomeFRER', FRERName)
            else:
                NodeID      = 'ns=4;s=' + excel_table['NodeID'][row]
            
            NodeID = NodeID.replace(' ', '')
            DB_data[str(sheet)][NodeID] = excel_table.iloc[row,1:].to_dict()

        # DB_NodeIDtoTag[NodeID] = excel_table['Tag'][row]
        # DB_TagtoNodeID[excel_table['Tag'][row]] = [NodeID]

    return DB_data

if __name__ == "__main__":

    DBdataConfig = ReadDBVarsFromExcel(Opts)

    Readings_JsonPath = os.path.join(Opts['ConfigDataPath'],Opts['json_OutputFileName_Readings'])
    # Saving the results
    with open(Readings_JsonPath, 'w') as file:
        json.dump(DBdataConfig, file)

    stop = 1